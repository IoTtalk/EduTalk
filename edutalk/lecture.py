import logging
import re
import os.path
import os
import datetime
import json
import time

from itertools import chain

from elasticsearch import Elasticsearch
from flask import Flask, Response, send_file, send_from_directory
from flask import Blueprint, render_template, session, abort, jsonify, request
from flask import render_template_string
from flask import redirect, url_for
from flask_login import current_user

from edutalk.config import config
from edutalk.models import Lecture, Template, LectureProject, get_pre_pid, set_pid, get_project_info
from edutalk.utils import login_required, teacher_required, json_err

from edutalk.exceptions import CCMAPIError
from edutalk.ag_ccmapi import devicemodel, devicefeature, device

from requests import get
from openpyxl import Workbook

app = Blueprint('lecture', __name__)
db = config.db
log = logging.getLogger('edutalk.lecture')

@app.route('/<int:id_>/bind',methods=['GET'], strict_slashes=False)
@login_required
def bind(id_):
    # print(current_user)
    # print(id_)
    # lecture_id = str(id_)+'abc' 
    # lecture = Lecture.query.get(id_)
    # x = LectureProject.get_by_lec_user(lecture, current_user)
    # pid = x.p_id
    # pre_pid = get_pre_pid()
    # re = set_pid(pid)
    # if pre_pid != 0:
    #     pre_logger_odo_id = get_project_info(pre_pid,current_user)
    #     device.unbind(pre_pid,pre_logger_odo_id)
    # logger_odo_id = get_project_info(pid,current_user)
    # logger_d_id = get_logger_info(pid,'d_id')
    # logger_d_id = get_logger_device_id(pid)
    # d_id_info = device.get(pid,logger_odo_id)
    # logger_d_id = d_id_info[0]['d_id']
    # device.bind(pid,logger_odo_id,logger_d_id)
    # print("testbind end")
    # return lecture_id+str(pid)+'abc'+str(logger_odo_id)+'abc'+str(logger_d_id)+'pre_pid='+str(pre_pid)
    pass
    return "ok"

@app.route('/<int:id_>/unbind',methods=['GET'], strict_slashes=False) 
@login_required
def unbind(id_):
    print(current_user)
    print(id_)
    lecture_id = str(id_)+'abc'
    lecture = Lecture.query.get(id_)
    x = LectureProject.get_by_lec_user(lecture, current_user)
    pid = x.p_id
    # logger_odo_id = get_project_info(pid,current_user)
#    logger_d_id = get_logger_info(pid,'d_id')
#    logger_d_id = get_logger_device_id(pid)
    # device.unbind(pid,logger_odo_id)
    # print("testunbind end")
    # return lecture_id+str(pid)+'abc'+str(logger_odo_id)+'abc'
    return "ok"


@app.route('/download_data',methods=['POST'], strict_slashes=False)
@login_required
def download_data():
    create_time = datetime.datetime.now()
    Sheet={}
    wb = Workbook()
    ws = wb.active
    worksheet_init(ws,3)
    sheet= wb['Sheet']

    filename="Oct03"
    timestamp=str(create_time.minute)+""+str(create_time.second)
    start_date = request.form['Ymd_s']
    start_hour = request.form['H_s']
    start_minute = request.form['i_s']
    start_second = request.form['s_s']
    start_unixtime = str(time_to_unixtime(start_date,start_hour,start_minute,start_second))
    start_string = start_date+start_hour+start_minute+start_second

    stop_date = request.form['Ymd_e']
    stop_hour = request.form['H_e']
    stop_minute = request.form['i_e']
    stop_second = request.form['s_e']
    stop_string = stop_date+stop_hour+stop_minute+stop_second
    stop_unixtime = str(time_to_unixtime(stop_date,stop_hour,stop_minute,stop_second))

#    start_unixtime = "1568623579962"
#    stop_unixtime  = "1568623580010"
    print("get into checkbox")
    export_sheet = 0
    for checkbox in 'Acceleration','Gyroscope','Magnetometer', 'Orientation':
        value = request.form.get(checkbox)
        if value:
            print("value="+value)
            ws = wb.create_sheet(value)
            worksheet_init(ws,3)
            query_es(value,start_unixtime,stop_unixtime,ws)
            export_sheet = 1
    for checkbox in 'Alcohol','Humidity','UV':
        value = request.form.get(checkbox)
        if value:
            print("value="+value)
            ws = wb.create_sheet(value)
            worksheet_init(ws,1)
            query_es(value,start_unixtime,stop_unixtime,ws)
            export_sheet = 1
    print("out of checkbox")
#    timestamp=start_time
#    filename="OCT"+timestamp
#    wb = Workbook()
#    ws = wb.active
#    worksheet_init(ws)
#    sheet= wb['Sheet']
#    timestamp=str(create_time.minute)+""+str(create_time.second)
#    sample=[12,13,14]
#    convert_time = 10
#    count=0
#    ws = wb.create_sheet("Gyr")
#    worksheet_init(ws)
#    worksheet_write(ws,start_unixtime,sample,convert_time,count)
#    worksheet_write(ws,timestamp,sample,convert_time,count)
    if export_sheet == 1:
        wb.remove(wb['Sheet'])
    wb.save("./docs/"+filename+'.xlsx')
    return send_file(os.path.join(os.getcwd()+"/docs", filename+'.xlsx'),
                     mimetype='text/csv',
                     attachment_filename=start_string+'&&'+stop_string+''+filename+'.xlsx',
                     as_attachment=True)

def query_es(df,time_start,time_stop,ws):
    ip = get('https://api.ipify.org').text
    es = Elasticsearch(hosts=ip)#"210.61.119.117"
#    df = "Acclogger"
#    time_start ="1568623579962"
#    time_stop ="1568623580010"
#    df = args.deviceFeature
#    time_start = args.timestart
#    time_stop = args.timestop
    convert_time = 10
    count=0
    print(df)
    if df == 'Acceleration' or df == 'Gyroscope' or df == 'Magnetometer' or df == 'Orientation':
        dim = 3
    if  df == 'Alcohol' or df == 'Humidity' or df == 'UV':
        dim = 1
    query_str ='{"_source":["sample","timestamp","to"],"query": {"bool": {"must": [{ "match": { "to": "'+df+'"}}],"filter":[{"range": {"timestamp": {"gte": "'+time_start+'" ,"lte":"'+time_stop+'"}}}]}}}'
    res = es.search(index="fluentd", body=query_str, from_=0, size=10000, pretty="true") # DON'T SET SIZE=100000
    #print(res['hits']['hits'])
    if (res['hits']['total']['value'] == 0):
        print("no sample, please try again!!")
    print("hits:",res['hits']['total']['value'])
    for hit in res['hits']['hits']:
        print (hit['_source']['timestamp'],hit['_source']['to'],hit['_source']['sample'])
        if count == 0:
            time_init = hit['_source']['timestamp']
        convert_time = hit['_source']['timestamp'] - time_init
        worksheet_write(ws,str(hit['_source']['timestamp']),hit['_source']['sample'],convert_time,count,dim)
        count = count+1
    return

def time_to_unixtime(date,hour,minute,second):
    #unixtime = int(time.mktime(time.strptime('YYYY-MM-DD HH:MM:SS', '%Y-%m-%d %H:%M:%S')))#e.g 2019-11-06 11:56:10 -> 1573012570
    GMT = 8 # 8 hour interval between taiwan(GMT+8) and GMT+0
    unixtime = int(time.mktime(time.strptime(''+date+' '+hour+':'+minute+':'+second+'', '%Y-%m-%d %H:%M:%S')))
    unixtime = unixtime-GMT*60*60
    return unixtime*1000

def worksheet_init(ws,dim):
    ws['A1'] = 'TimeStamp'
    ws.column_dimensions['A'].width = 25.0
    alpha = 'B'
    for i in range(1,dim+1):
        column = alpha+'1'
        column_content = 'x'+str(i)
        ws[column] = column_content
        ws.column_dimensions[alpha].width = 20.0
        alpha = chr(ord(alpha) + 1)
    ws[alpha+'1'] = 'Interarrival time interval (ms)'
    #µs->ms
    ws.column_dimensions[alpha].width = 30.0
    #ws['B1'] = 'x1'
    #ws['C1'] = 'x2'
    #ws['D1'] = 'x3'
    #ws['E1'] = 'Interarrival time interval (µs)'
    #ws.column_dimensions['A'].width = 25.0
    #ws.column_dimensions['B'].width = 20.0
    #ws.column_dimensions['C'].width = 20.0
    #ws.column_dimensions['D'].width = 20.0
    #ws.column_dimensions['E'].width = 30.0
    return

def worksheet_write(ws,timestamp,sample,convert_time,count,dim):
    row = count + 2
    GMT = 8 # 8 hour interval between taiwan(GMT+8) and GMT+0
    unixtime = int(timestamp)+GMT*60*60*1000
    front_time = unixtime/1000
    millisecond = unixtime%1000
    readable_timestamp = datetime.datetime.fromtimestamp(front_time).strftime('%Y-%m-%d %H:%M:%S.')+str(millisecond)
#    L = np.array(sample)
#    dim = L.shape[0]
#    dim = 3
    alpha = 'B'
    ws['A'+ str(row)] = readable_timestamp
    for i in range(1,dim+1):
        ws[alpha+ str(row)] = float(sample[i-1])
        alpha = chr(ord(alpha) + 1)
    #ws['A'+ str(row)] = timestamp
    #ws['B'+ str(row)] = str(sample[0])
    #ws['C'+ str(row)] = str(sample[1])
    #ws['D'+ str(row)] = str(sample[2])
    ws[alpha+ str(row)] = convert_time
    return

@app.route('/<int:id_>', methods=['GET'], strict_slashes=False)
@login_required
def detail(id_):
    lec = Lecture.query.get(id_)
    if lec is None:
        abort(404)

    templates = tuple(chain(*Template.query.values(Template.dm)))
    lesson_data = Lecture.list_()

    df_list = tuple(map(
        lambda x: x.get('df_name'),
        devicemodel.get(lec.odm)['df_list']))
    LectureProject.get_or_create(current_user, lec)

    return render_template('tutorial.html',
                           lecture=lec,
                           t_list=[],
                           lesson_data=lesson_data,
                           template_list=templates,
                           df_list=df_list,
                           token=current_user.token)


@app.route('/', methods=['GET'], strict_slashes=False)
@login_required
def index():  # place holder for ``url_for``
    ...


@app.route('/<int:id_>/rename', methods=['POST'], strict_slashes=False)
@login_required
@teacher_required
def rename(id_):
    '''
    Request::

        {
            'name': 'my new name',
        }
    '''
    name = request.json.get('name')
    if not name:
        return json_err('Lecture name cannot be empty'), 400
    if Lecture.isexist(name):
        return json_err('Lecture name already exists'), 400

    l = Lecture.query.get(id_)
    l.name = name
    db.session.commit()
    return jsonify({'state': 'ok'})


@app.route('/<int:id_>/url', methods=['POST'], strict_slashes=False)
@login_required
@teacher_required
def update_url(id_):
    '''
    Request::

        {
            url: 'http://.../new_url',
        }
    '''
    url = request.json.get('url')
    if not url:
        return json_err('url cannot be empty'), 400

    l = Lecture.query.get(id_)
    l.url = url
    db.session.commit()
    return jsonify({'state': 'ok'})


@app.route('/reorder', methods=['POST'], strict_slashes=False)
@login_required
@teacher_required
def reorder():
    '''
    Request::

        {
            'order': [id1, id2, id3, ...],
        }
    '''
    o = request.json.get('order')
    if not o:
        return json_err('Order is empty'), 400

    for idx, id in enumerate(o):
        id = str(id)
        l = Lecture.query.get(id)
        if l is None:
            return json_err('Lecture {} not found'.format(id)), 400
        l.idx = idx

    db.session.commit()
    return jsonify({'state': 'ok'})


@app.route('/<int:id_>/delete', strict_slashes=False)
@login_required
@teacher_required
def delete(id_):
    l = Lecture.query.get(id_)
    if l is None:
        return json_err('Lecture {} not found'.format(id_)), 404

    for x in l.lecture_projects:
        x.delete()


    # delete input/output device model

    devicemodel.delete(l.idm)
    devicemodel.delete(l.odm)

    '''
    delete input/output device feature which name starts with 'odm_'
    '''
    for j in l.joins:
        for df in j:
            if type(df) == str and df.startswith(l.odm+'_'):
                devicefeature.delete(df)

    db.session.delete(l)
    db.session.commit()

    return redirect(url_for('index'))


@app.route('/create', methods=['GET'], strict_slashes=False)
@login_required
@teacher_required
def create_page():
    t_list = [{'name': 'New', 'df_list': []}] + [
        {
            'name': dm,
            'df_list': list(map(
                lambda x: x.get('df_name'),
                devicemodel.get(dm)['df_list']))
        }
        for dm in chain(*Template.query.values(Template.dm)) if dm != 'New'
    ]

    return render_template('tutorial.html',
                           lecture=None,
                           lesson_data=Lecture.list_(),
                           t_list=t_list,
                           df_list=[])


@app.route('/create', methods=['PUT'], strict_slashes=False)
@login_required
@teacher_required
def create():
    '''
    Request::
{
            'name': 'Magic',
            'url': 'HackMD URL',  // optional
            'odm': {
                'name': 'FooModel',
                'dfs': [{'name': odf_name}, ...],
            }
            'idm': {
                'name': 'BarModel',
                'dfs' [{
                    'name': idf_name,
                    'min': min,  // optional
                    'max:' max,  // optional
                }, ...],
            },
            'joins': [[idf, odf], [idf, odf]],
            'code': ...,
        }
    '''
    log.debug('create lecture payload: %s', request.json)

    for f in ('name', 'odm', 'idm', 'joins', 'code'):
        if f not in request.json:
            return json_err('field `{}` is required'.format(f)), 400
    name = request.json['name']
    if not name:
        return json_err('Lecture name is required', type='lecture'), 400
    elif Lecture.isexist(name):
        return json_err('duplicated lecture name', type='lecture'), 400
    url = request.json.get('url', '')

    # checking all fields first, then invoke ccmapi to ``create`` stuff
    for dm in ('odm', 'idm'):
        x = request.json.get(dm)
        if not x.get('name'):
            return json_err('Program name is required', type='dm'), 400
        if 'dfs' not in x:
            msg = 'device feature list is for {} is required'.format(dm)
            return json_err(msg, type='df'), 400
    # check min, max is not empty if this field exists
    idfs = request.json['idm']['dfs']
    for df_info in idfs:
        # check if parameter value is reasonable
        if all(limit in df_info for limit in ('min', 'max', 'default')):
            _min, _max, _default = df_info['min'], df_info['max'], df_info['default']
            for p in ('min', 'max', 'default'):
                if df_info[p] == "":
                    msg = '{} cannot be empty'.format(p)
                    return json_err(msg, type='df_parameter'), 400
            if not _min <= _max:
                msg = 'min value should be <= max value'
                return json_err(msg, type='df_parameter'), 400
            if _default not in range(_min, _max+1):
                msg = 'default value should be between min and max'
                return json_err(msg, type='df_parameter'), 400

    # check odm_name is eng+number only
    odm_name = request.json['odm']['name']
    if re.match('[a-zA-Z_][a-zA-Z0-9_]*', odm_name) is None:
        return json_err('Invalid program name, english alphabet or number only', type='dm'), 400
    # check odm is non-exists
    try:
        status = devicemodel.get(odm_name) #fixd bugs
        if not status:
            raise CCMAPIError
        return json_err('Program name already used by other lecture', type='dm'), 400
    except CCMAPIError as e:
        ...
    # create df of dm
    for dm, typ in (('odm', 'output'), ('idm', 'input')):
        x = request.json.get(dm)
        if not x['dfs']:
            return json_err('Feature list cannot be empty', type='df'), 400
        for df in x['dfs']:
            devicefeature.get_or_create(re.sub(r'_', r'-', df['name']), 
                                        typ, 
                                        [{'param_type': 'float', 'min': 0, 'max': 10,}]
            )



    # create dm
    for dm in ('idm', 'odm'):
        dm = request.json[dm]
        devicemodel.create(dm['name'], [{
            'key': re.sub(r'_', r'-', x['name']) if type(x) is dict else x,
            'parameter': [{
                'param_type': 'float',
                'min': x['min'],
                'max': x['max'],
            }] if type(x) is dict and all(limit in x for limit in ('min','max')) else [{}]
        } for x in dm['dfs']])

    odm = request.json['odm']
    sensor_odfs = ['Acceleration_O', 'Gyroscope_O', 'Orientation_O'] 
    odfs = list(map(lambda x: [x['name'], ['', '', '']] if x['name'] in sensor_odfs else [x['name'],['']], odm['dfs']))
    sm_odfs = [x['name'] for x in odm['dfs'] if re.search('_O$', x['name'])]

    df_default_values = {odm['dfs'][idx]['name']: x['default'] if 'default' in x else 0 for idx, x in enumerate(request.json['idm']['dfs'])}
    odm_temp = request.json['code']
    temp = Template.query.filter_by(dm=odm_temp).first()
    print("*&^&*%&^%&^%")
    vp_code = render_template_string(temp.code,
                                     dm_name=odm['name'],
                                     odf_list=odfs,
                                     sm_odf_list=sm_odfs,
                                     df_default_values=df_default_values,
                                     trim_blocks=True,
                                     lstrip_blocks=True)

    lec = Lecture(name=name, idx=len(Lecture.query.all()) + 1, url=url,
                  idm=request.json['idm']['name'],
                  odm=request.json['odm']['name'],
                  joins=request.json['joins'],
                  code=vp_code)
    db.session.add(lec)
    db.session.commit()

    return jsonify({'state': 'ok', 'url': url_for('lecture.detail', id_=lec.id)})


@app.app_template_filter()
def todf_list(x):
    return '[{}]'.format(', '.join('[{}, {}]'.format(a[0], repr(a[1])) for a in x))
